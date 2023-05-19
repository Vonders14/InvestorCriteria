import openpyxl

# File path
file_path = r"C:\Users\dhoffmann\Documents\MULTIFAMILY_RE Criteria.xlsx"

# Load workbook
workbook = openpyxl.load_workbook(file_path)

# Select the first sheet in the workbook
sheet = workbook.worksheets[0]

# Define keywords to search for
keywords = [
    "AL", "Alabama", "AK", "Alaska", "AZ", "Arizona", "AR", "Arkansas", "CA", "California", "CO", "Colorado",
    "CT", "Connecticut", "DE", "Delaware", "FL", "Florida", "GA", "Georgia", "HI", "Hawaii", "ID", "Idaho",
    "IL", "Illinois", "IN", "Indiana", "IA", "Iowa", "KS", "Kansas", "KY", "Kentucky", "LA", "Louisiana",
    "ME", "Maine", "MD", "Maryland", "MA", "Massachusetts", "MI", "Michigan", "MN", "Minnesota", "MS", "Mississippi",
    "MO", "Missouri", "MT", "Montana", "NE", "Nebraska", "NV", "Nevada", "NH", "New Hampshire", "NJ", "New Jersey",
    "NM", "New Mexico", "NY", "New York", "NC", "North Carolina", "ND", "North Dakota", "OH", "Ohio", "OK", "Oklahoma",
    "OR", "Oregon", "PA", "Pennsylvania", "RI", "Rhode Island", "SC", "South Carolina", "SD", "South Dakota",
    "TN", "Tennessee", "TX", "Texas", "UT", "Utah", "VT", "Vermont", "VA", "Virginia", "WA", "Washington",
    "WV", "West Virginia", "WI", "Wisconsin", "WY", "Wyoming"
]

# Add "Target States" column header in column AW if not already present
if sheet["AW1"].value != "Target States":
    sheet["AW1"] = "Target States"

# Iterate through each row in the sheet
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    target_states = []  # List to store found keywords
    for cell in row:
        if cell is not None:
            cell_value = str(cell)
            for keyword in keywords:
                if keyword.lower() in cell_value.lower() or keyword.lower() == cell_value[:2].lower():
                    target_states.append(keyword)  # Add found keyword to target states list

    # Set the corresponding cell in column AW with the results
    target_states_cell = sheet.cell(row=i, column=sheet.max_column)
    target_states_cell.value = ", ".join(target_states)   # Join the target states list into a string

# Save the workbook
workbook.save(file_path)
