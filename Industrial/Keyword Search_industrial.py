import openpyxl
import os
import re

# File paths
source_file_path = r"C:\Users\dhoffmann\Documents\Book2-property retest.xlsx"
destination_directory = "C:/Users/dhoffmann/Documents"
target_workbook_name = "mf retest.xlsx"

# Load source workbook
source_workbook = openpyxl.load_workbook(source_file_path)

# Select the specific sheet by index
target_sheet_index = 0  # Index of the third sheet (zero-based index)
source_sheet = source_workbook.worksheets[target_sheet_index]

# Create a new workbook
destination_workbook = openpyxl.Workbook()
destination_sheet = destination_workbook.active

# Copy header row from source sheet to destination sheet
header_row = []
for cell in source_sheet[1]:
    header_row.append(cell.value)
destination_sheet.append(header_row)

# Define keywords to search for
keywords = [
    "industrial"
   
]

# Iterate through each row in the source sheet
for row in source_sheet.iter_rows(min_row=2, values_only=True):
    for cell in row:
        if cell is not None and any(
                re.search(r"\b{}\b".format(keyword), str(cell), re.IGNORECASE) for keyword in keywords):
            destination_sheet.append(row)
            break

# Save the destination workbook to the specified path
destination_file_path = os.path.join(destination_directory, target_workbook_name)
destination_workbook.save(destination_file_path)
print("Workbook saved successfully.")

# Print the number of records found
num_records = destination_sheet.max_row - 1  # Subtract 1 to exclude the header row
print(f"Number of records found: {num_records}")