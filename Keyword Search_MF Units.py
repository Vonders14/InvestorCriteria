import openpyxl
import re

# File path
source_file_path = r"C:\Users\dhoffmann\Documents\MULTIFAMILY_RE Criteria.xlsx"

# Load the workbook
source_workbook = openpyxl.load_workbook(source_file_path)

# Select the specific sheet by index or name
target_sheet_index = 0  # Replace with the index or name of the desired sheet
source_sheet = source_workbook.worksheets[target_sheet_index]

# Define the range of columns (A to AT) to search for keywords
start_column = 'A'
end_column = 'AT'

# Define the keywords to search for
keywords = ["unit", "units", "Unit", "Units"]

# Define the target column for storing the extracted words
target_column = 'AU'

# Get the column index of the start and end columns
start_column_index = openpyxl.utils.column_index_from_string(start_column)
end_column_index = openpyxl.utils.column_index_from_string(end_column)

# Create a regex pattern for exact word match and non-case sensitive search
pattern = r"\b(" + "|".join(keywords) + r")\b"

# Remove the current text in the target column
for cell in source_sheet[target_column]:
    cell.value = None

# Iterate through all columns, including hidden columns
for column in source_sheet.columns:
    column_letter = column[0].column_letter
    if start_column <= column_letter <= end_column:
        for cell in column:
            if cell.row == 1:
                continue

            if cell.value is not None and isinstance(cell.value, str):
                # Use regex pattern for exact word match and non-case sensitive search
                matches = re.finditer(pattern, cell.value, flags=re.IGNORECASE)
                extracted_words = []
                for match in matches:
                    start_index = max(match.start() - 30, 0)
                    end_index = match.end() + 30
                    text_preview = cell.value[start_index:end_index]
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    extracted_words.append(text_preview)

                result = ", ".join(extracted_words)
                result_cell = source_sheet.cell(row=cell.row, column=openpyxl.utils.column_index_from_string(target_column))
                result_cell.value = result

# Save the updated workbook
source_workbook.save(source_file_path)
