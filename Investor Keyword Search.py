import openpyxl
import os

# File paths
source_file_path = r"C:\Users\dhoffmann\Downloads\Apollo RE Investor Criteria_MASTER - COPY_bad text removed.xlsx"
destination_directory = "C:/Users/dhoffmann/Documents"
target_workbook_name = "MULTIFAMILY_RE Criteria v3.xlsx"

# Load source workbook
source_workbook = openpyxl.load_workbook(source_file_path)

# Select the specific sheet by index
target_sheet_index = 2  # Index of the third sheet (zero-based index)
source_sheet = source_workbook.worksheets[target_sheet_index]

# Create a new workbook
destination_workbook = openpyxl.Workbook()
destination_sheet = destination_workbook.active

# Copy header row from source sheet to destination sheet
header_row = []
for cell in source_sheet[1]:
    header_row.append(cell.value)
destination_sheet.append(header_row)

# Define keywords to search for
keywords = ["apartment",
"apartments",
"multifamily",
"multi-family"]

# Iterate through each row in the source sheet
for row in source_sheet.iter_rows(min_row=2, values_only=True):
    if any(cell is not None and keyword.lower() in str(cell).lower() for keyword in keywords for cell in row):
        destination_sheet.append(row)

# Save the destination workbook to the specified path
destination_file_path = os.path.join(destination_directory, target_workbook_name)
destination_workbook.save(destination_file_path)
print("Workbook saved successfully.")
